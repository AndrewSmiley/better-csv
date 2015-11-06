__author__ = 'pridemai'
import traceback
import pandas as pd
from openpyxl import *
import sys;
reload(sys);
sys.setdefaultencoding("ascii")

def load_excel_with_pandas(filename):

    print "reading data from %s" %(filename)
    data = pd.read_excel(filename)
    data.fillna('',inplace=True)
    print "reading data from %s" %(filename)
    print "loading data from %s" %(filename)
    # Output the number of rows
    indexer = 0
    data_list = []
    names = data.keys()
    #preload the header columns
    data_list.append(list(x for x in names))
    while indexer < len (data[names[0]]):
        print "Working row %s of %s in file %s" % (indexer, len(data[names[0]]), filename)
        _new = []
        for n in names:
            _new.append(data[n][indexer])
        data_list.append(_new)
        indexer = indexer+1

    print "Finished loading file %s" %(filename)

    return data_list

def write_to_excel(data,filename):
    """
    write a multidimensional list to excel
    :param data: the list of rows as lists
    :return: None
    """
    print "Saving file %s" %(filename)
    row = 0
    w = Workbook()
    sheet = w.create_sheet()
    while row < len(data):
        column = 0
        while column < len(data[row]):
            sheet.cell(row=row+1,column=column+1).value = data[row][column] if data[row][column] != '' and data[row][column] != None else ' '
            column = column + 1

        row = row + 1

    w.save('results/%s'%filename)


def write_to_excel_redux(data,filename):

    master_copy = load_workbook(filename = "data/%s" %(filename))
    master_sheet = master_copy.get_sheet_by_name(master_copy.get_sheet_names()[0])

    row = master_sheet.min_row
    while row <= master_sheet.max_row:

        column = master_sheet.min_column
        while column <= master_sheet.max_column:
            try:
                master_sheet.cell(row=row,column=column).value = data[row-1][column-1]
            except:
                print "An error occurred.\rAttempted to update row %s column %s\r" \
                      "Data row contained %s objects\r\r%s\r\r" % (row,column, len(data[row-1]), data[row-1])
                traceback.print_exc()
            column = column+1
        row = row+1

    master_copy.save("results/%s"%(filename))



