__author__ = 'pridemai'
from models import File,ColumnMapping,SearchColumn
from better_csv_web.settings import *
import traceback
import time
from os import listdir
from os.path import isfile, join
from parse_functions import *
import time
from openpyxl import *
DATA_DIR = "/data/"
MASTERS_DIR = "/results/"

def write_csv(filepath,rows):
    hs = open(filepath,"w+")
    for mline in rows:
        hs.write(",".join(mline)+"\r")

    hs.close()
def get_all_source_file_names():
    return File.objects.filter(is_master=False)

def get_all_master_file_names():
    return File.objects.filter(is_master=True)

def upload_file(request, is_master):
    file = File()
    file.filename = request.FILES['csv'].name
    file.is_master = is_master
    file.save()
    #write the file to the disk
    with open(BASE_DIR+DATA_DIR+request.FILES['csv'].name, 'wb+') as destination:
        for chunk in request.FILES['csv'].chunks():
            destination.write(chunk)


def rename_file(fname, new_filename):
    try:
        file = File.objects.get(filename=fname)
        file.filename = new_filename
        file.save()
        return {"result":True}
    except:
        return {"result":False, "message": traceback.print_exc()}

def execute(request):
    start = time.time()
    messages=[]
    # exact_matches = True if 'exact_matches' in request.POST else False
    exact_matches = True if request.POST.getlist('exact_matches')[0] == "True" else False
    for master in request.POST.getlist('master_files[]'):
        if ".csv" in master :
            master_copy = BetterCSV().get_lists(BetterCSV().get_lines(open(BASE_DIR + DATA_DIR + master).read()))

            for file in request.POST.getlist('source_files[]'):
                if ".csv" in file:
                    print "Processing file"+file
                    data_file = BetterCSV().get_lists(BetterCSV().get_lines(open(BASE_DIR + DATA_DIR + file).read()))
                    # load the search columns
                    master_search_columns = []
                    for msc in SearchColumn.objects.filter(file=File.objects.get(filename=master)):
                        master_search_columns.append(msc.column_id - 1)
                    data_search_columns = []
                    for dsc in SearchColumn.objects.filter(file=File.objects.get(filename=file)):
                        data_search_columns.append(dsc.column_id - 1)
                        # load the mappings
                    mappings = {}
                    for mapping in ColumnMapping.objects.filter(master_file=File.objects.get(filename=master),
                                                                source_file=File.objects.get(filename=file)):
                        mappings[mapping.master_column_id - 1] = mapping.source_column_id - 1
                        # actually execute and update
                    results = iterate(master_copy, data_file, master_search_columns, data_search_columns, mappings, file,exact_matches)
                    master_copy = results['data']
                    messages.append(results['message'])
                elif ".xls" in file or ".xlsx" in file:
                    data_wb = load_workbook(filename = 'data/%s'%(file))
                    data_sheet = data_wb.get_sheet_by_name(data_wb.get_sheet_names()[0])

                    try:
                        result = iterate_master_csv_data_excel(master_copy,list(x for x in [[data_sheet.cell(row=y, column=z)  for z in range(data_sheet.min_column, data_sheet.max_column)] for y in range(data_sheet.min_row, data_sheet.max_row)]),
                                                                list(x.column_id for x in SearchColumn.objects.filter(file=File.objects.get(filename=master))),
                                                                list(x.column_id for x in SearchColumn.objects.filter(file=File.objects.get(filename=file))),
                                                                dict((str(x.master_column_id), str(x.source_column_id)) for x in ColumnMapping.objects.filter(master_file=File.objects.get(filename=master),
                                                                source_file=File.objects.get(filename=file))), file, exact_matches)
                    except:
                        traceback.print_exc()
                    master_copy = result['data']
                    messages.append(result['message'])

                write_csv(str(BASE_DIR+MASTERS_DIR+master.split(".")[0]+"_"+str(time.strftime("%d%m%Y"))+".csv"), master_copy)
        elif ".xls" in master or ".xlsx" in master:
            master_copy = load_workbook(filename = "data/%s" %(master))
            master_sheet = master_copy.get_sheet_by_name(master_copy.get_sheet_names()[0])
            #iterate over the source files
            for file in request.POST.getlist('source_files[]'):
                #check if it's excel or not
                if ".xls" in file or ".xlsx" in file:
                    data_wb = load_workbook(filename = 'data/%s'%(file))
                    data_sheet = data_wb.get_sheet_by_name(data_wb.get_sheet_names()[0])
                    dict((x.master_column_id, x.source_column_id) for x in ColumnMapping.objects.filter(master_file=File.objects.get(filename=master),
                                                            source_file=File.objects.get(filename=file)))
                    result = iterate_excel(master_sheet,list(x for x in [[data_sheet.cell(row=y, column=z)  for z in range(data_sheet.min_column, data_sheet.max_column)] for y in range(data_sheet.min_row, data_sheet.max_row)]),
                                  list(x.column_id for x in SearchColumn.objects.filter(file=File.objects.get(filename=master))),
                                  list(x.column_id for x in SearchColumn.objects.filter(file=File.objects.get(filename=file))),
                                  dict((str(x.master_column_id), str(x.source_column_id)) for x in ColumnMapping.objects.filter(master_file=File.objects.get(filename=master),
                                                            source_file=File.objects.get(filename=file))), file, exact_matches)
                    master_sheet = result['data']

                    messages.append(result['message'])
                    #ok so if it's an excel file use the iterate function that we had before
                    pass
                #otherwise check for csv
                elif ".csv" in file:

                    data_copy = BetterCSV().get_lists(BetterCSV().get_lines(open(BASE_DIR+DATA_DIR+file).read()))
                    result = iterate_master_excel_data_csv(master_sheet,data_copy,
                                                  list(x.column_id for x in SearchColumn.objects.filter(file=File.objects.get(filename=master))),
                                                  list(x.column_id for x in SearchColumn.objects.filter(file=File.objects.get(filename=file))),
                                                   dict((str(x.master_column_id), str(x.source_column_id)) for x in ColumnMapping.objects.filter(master_file=File.objects.get(filename=master),
                                                            source_file=File.objects.get(filename=file))),
                                                  file,exact_matches)
                    master_sheet = result['data']
                    messages.append(result['message'])


                #otherwise we want to just move to the next file cause something is fucky
                else:
                    continue
            try:
                master_copy._add_sheet(master_sheet)
                master_copy.save("results/"+master)

            except:
                print "saving file"
                master_copy.save("results/"+master)

        else:
            continue
    return {"messages": messages, "runtime":time.time()-start}




def update_row(master_row, data_row, column_mapping):
    for key,value in column_mapping.iteritems():

        master_row[key]=data_row[value] if len(data_row[value]) > 0 and data_row[value] != "0" else master_row[key]
    return master_row

def iterate(master_copy,data_copy, master_search_columns, data_search_columns, column_mapping, filname="N/A",exact_matches=False):

    better_csv = BetterCSV()
    new_master = []
    found_count = 0
    for line in master_copy:
        for m in master_search_columns:
            found = False
            for d in data_search_columns:
                data_copy = sorted(data_copy, key=lambda x: x[d], reverse=False)
                results = binary_search(line, data_copy,m, d,exact_matches)
                found = results["result"]
                if found:
                    line = update_row(line, data_copy[results["index"]], column_mapping)
                    found_count = found_count +1
                    break
            if found:
                break
        new_master.append(line)

        # master_args=[]
        # for column in master_search_columns:
        #     master_args.append(line[column])
        # for d in data_copy:
        #     data_args =[]
        #     for column in data_search_columns:
        #         data_args.append(d[column])
        #     if better_csv.search(master_args, data_args) :
        #         line = update_row(line, d, column_mapping)
        #         found_count = found_count + 1You
        #         break
        # new_master.append(line)
    return {"data":new_master, "message": "%s count: %s" % (filname, found_count) }


def iterate_master_excel_data_csv(master_copy,data_copy, master_search_columns, data_search_columns, column_mapping, filname="N/A",exact_matches=False):
    master_copy_row = master_copy.min_row
    found_count=0

    while master_copy_row <= master_copy.max_row:
        try:
            for m in master_search_columns:
                found = False
                # print "Processing row %s in file %s" % (master_copy.cell(row=master_copy_row, column=m).value, filname)
                for d in data_search_columns:
                    print "Processing row %s in file %s %s" % (master_copy.cell(row=master_copy_row, column=m).value, filname,d)
                    data_copy = sorted(data_copy, key=lambda x: x[d-1], reverse=False)
                    results = basic_binary_search_with_searchkey(str(master_copy.cell(row=master_copy_row, column=m).value), data_copy,d-1, exact_matches)
                    found = results['result']
                    if found:
                        found_count = found_count +1
                        master_copy_row = master_copy_row + 1
                        for x,y in column_mapping.iteritems():
                            if data_copy[int(results['index'])][int(y)-1] != '' and data_copy[int(results['index'])][int(y)-1] != None:
                                master_copy.cell(row=master_copy_row, column=int(x)).value=data_copy[int(results['index'])][int(y)-1]
                        break
                if found:
                    break
            master_copy_row = master_copy_row + 1
        except:
            traceback.print_exc()
            master_copy_row = master_copy_row + 1
    return {"data": master_copy, "message":"%s count %s" % (filname, found_count)}

def iterate_master_csv_data_excel(master_copy,data_copy, master_search_columns, data_search_columns, column_mapping, filname="N/A",exact_matches=False):
    # data_copy_row = data_copy.min_row
    #
    # while data_copy_row <= data_copy.max_row:
    #     for m in master_search_columns:
    #         for d in data_search_columns:
    new_master_copy = []
    found_count = 0
    for master_line in master_copy:
        found = False
        for m in master_search_columns:
            for d in data_search_columns:
                print "Searching for matches with %s" % (master_line[m-1])
                results = basic_binary_search_with_searchkey(master_line[m-1],
                                                             sorted(data_copy,key=lambda x:x[d-1].value,reverse=False),
                                                            d-1, exact_matches)
                found = results['result']
                if found:
                    found_count = found_count+1
                    for x,y in column_mapping.iteritems():
                        master_line[int(x)-1] = data_copy.cell(row=results['index'], column=int(y)).value

                    break
            if found:
                break


        new_master_copy.append(master_line)

    return {"data":new_master_copy,"message":"%s count %s" % (filname, found_count)}



    pass

def iterate_excel(master_copy,data_copy, master_search_columns, data_search_columns, column_mapping, filname="N/A",exact_matches=False):
    """
    Function to iterate over two excel files, find matches and update data
    :param master_copy:  an openpyxl worksheet
    :param data_copy: an openpyxl worksheet
    :param master_search_columns:
    :param data_search_columns:
    :param column_mapping:
    :param filname:
    :return:
    """
    for x,y in column_mapping.iteritems():
        print "%s, %s" %(x,y)
    master_copy_row = master_copy.min_row
    found_count=0

    while master_copy_row <= master_copy.max_row:
        # data_copy_row = data_copy.min_row

        # master_args = list(master_copy.cell(row=master_copy_row, column=x).value for x in master_search_columns)

        for m in master_search_columns:
            try:
                found = False
                for d in data_search_columns:
                    d = d-1
                    data_copy = sorted(data_copy, key=lambda x: x[d].value, reverse=False)
                    # searchable_data_list = list(x[d] for x in data_copy)
                    result = excel_binary_search(str(master_copy.cell(row=master_copy_row, column=m).value), data_copy, d,exact_matches)
                    found = result['result']
                    if found:
                        # print "Match Found"
                        found_count = found_count+1
                        for x,y in column_mapping.iteritems():
                            print "Updating row %s %s:%s : %s=>%s" %(master_copy_row,x,y,master_copy.cell(row=master_copy_row, column=int(x)).value,data_copy[int(result['index'])][int(y)-1].value)
                            master_copy.cell(row=master_copy_row, column=int(x)).value = data_copy[int(result['index'])][int(y)-1].value
                            master_copy_row = master_copy_row + 1
                        break
            except Exception,e:
                traceback.print_exc()

            if found:
                break
        master_copy_row = master_copy_row +1





    return {"data":master_copy, "message": "%s count: %s" % (filname, found_count) }
def get_result_files():

    return [ f for f in listdir(BASE_DIR+MASTERS_DIR) if isfile(join(BASE_DIR+MASTERS_DIR,f)) ]

def get_all_files():
    return [ f for f in listdir(BASE_DIR+MASTERS_DIR) if isfile(join(BASE_DIR+MASTERS_DIR,f)) ] + [ f for f in listdir(BASE_DIR+DATA_DIR) if isfile(join(BASE_DIR+DATA_DIR,f)) ]

def get_file(filepath, filename):
    return open(filepath+filename)

def get_files_in_folder(filepath):
    return [ f for f in listdir(filepath) if isfile(join(filepath,f)) ]

def array_to_html_tr(array):
    return "<tr>%s</tr>" % (''.join(list("<td>%s</td>" % (x) for x in array)))

def search_in_files(request):
    search_results = []
    for file in request.POST.getlist('search_files[]'):
        rows =BetterCSV().get_lists(BetterCSV().get_lines(get_file("%s/%s/"%(BASE_DIR,request.POST['folder']), file).read()))
        start = time.time()
    #   the binary search way
        count = 0
        for row in rows:

            # row = sorted(row, key=lambda x: x[i], reverse=False)
            try:
                if basic_binary_search(request.POST['search_term'], sorted(row)):
                    html = "<table style=\"width: 100%%\" border=\"2\">%s%s</table><br><br>" % (array_to_html_tr(rows[0]),array_to_html_tr(row).replace(request.POST['search_term'],"<mark>%s</mark>" % (request.POST['search_term'])))
                    first_part = "<h3>Result found in row %s of file %s</h3><br>" % (count+1, file)
                    search_results.append(first_part+html)
                    #search_results.append(("Result found in row %s of file %s<br> <table style=\"width: 100%\" border=\"2\">%s</table><br><br>" % (str(count), file, array_to_html_tr(row))).decode('utf-8'))
                count = count+1
            except Exception, e:
                count = count+1
                print str(e)
                continue
        print str("%s seconds to search file %s" % (str(time.time()-start), file))

    return search_results


def excel_to_csv(sheet):
    data_list=[]
    row = sheet.min_row
    while row <= sheet.max_row:
        # print str(row)
        row_data = []
        col = sheet.min_column
        while col <= sheet.max_column:
            print sheet.cell(row=row, column=col).value
            try:

                row_data.append(str(sheet.cell(row=row, column=col).value))
                col=col+1
            except Exception,e:
                print str(e)
                col=col+1

        row = row+1
        data_list.append(row_data)


    return data_list