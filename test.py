from openpyxl import *
from parse_functions import *
import threading
import traceback
import sys;
reload(sys);
sys.setdefaultencoding("utf8")

for x,y in {'22':'22', '23':'35'}.iteritems():
    print "%s %s" % (x,y)
# master_book = load_workbook(filename='data/Copy of leah_22102015 colorcodedweightbd.xlsx')
# master_sheet = master_book.get_sheet_by_name(master_book.get_sheet_names()[0])
# new_sheet = master_sheet

# new_workbook = Workbook()
# new_workbook._add_sheet(new_sheet,0)
# new_workbook.save("tester.xlsx")
# import pandas as pd
# print "new revision"
# # Read the file
# names =['Part Number','Nomenclature','Demanding Org','Engine Code','Engine Program','Supplier',"2015 (Oct-Dec)",'2016','2017','2018','2019','2020','2021','2022','2023','2024']
# data = pd.read_csv("data/forecast.csv", names=names)
# # Output the number of rows
# indexer = 0
# data_list = []
# while indexer < len (data['Part Number']):
#     # print "Wokring row %s" % (indexer)
#     _new = []
#     for n in names:
#         _new.append(str(data[n][indexer]).encode('utf-8','ignore'))
#
#     data_list.append(_new)
#     indexer = indexer+1
# data = data_list
# print("Total rows: {0}".format(len(data)))
book2 = load_workbook(filename='data/11-2-2015 leah.xlsx')
# print "loaded workbook"

data_sheet = book2.get_sheet_by_name(book2.get_sheet_names()[0])
master_book = load_workbook(filename='updates.xlsx')
master_sheet = master_book.get_sheet_by_name(master_book.get_sheet_names()[0])
# # data_sheet = BetterCSV().get_lists(BetterCSV().get_lines(open("data/forecast.csv").read()))
# print "about to get data"
# # sorted_data = list(x for x in
# #                    [[data_sheet.cell(row=y, column=z) for z in range(data_sheet.min_column, data_sheet.max_column)] for
# #                     y in range(data_sheet.min_row, data_sheet.max_row)])
# #
# # for x in sorted(sorted_data, key=lambda x: x[1].value, reverse=False)[0]:
# #     print x.value
# for x in data:
#     print x[0]
print "finished loading the data"
master_search_columns = [4]
data_search_columns = [4]
master_row = master_sheet.min_row
found_count = 0
column_mappings={'21':21}
data_sheet = list(x for x in [[data_sheet.cell(row=y, column=z) for z in range(data_sheet.min_column, data_sheet.max_column)] for y in range(data_sheet.min_row, data_sheet.max_row)])
while master_row <= master_sheet.max_row:

    for m in master_search_columns:
        found = False
        for d in data_search_columns:
            try:
                data_sheet= sorted(data_sheet, key=lambda x: x[d-1].value, reverse=False)
                # res = basic_binary_search_with_searchkey(str(master_sheet.cell(row=master_row, column=m).value).encode('utf-8','ignore'), data, d, True)
                # sorted_data = sorted(sorted_data, key=lambda x: x[d].value, reverse=False)
                # print "processing part %s" % (master_sheet.cell(row=master_row, column=m ).value)
                res = excel_binary_search(str(master_sheet.cell(row=master_row, column=m).value), data_sheet, d-1, False)
                if res['result'] == True:
                    # print "match found: %s == %s" % (master_sheet.cell(row=master_row, column=m).value, sorted_data[int(res['index'])][d].value)
                    found = True
                    found_count = found_count + 1
                    for x,y in column_mappings.iteritems():
                        # print "Updating column %s=>%s" % (x,y)
                        if data_sheet[int(res['index'])][int(y)-1] != '' and data_sheet[int(res['index'])][int(y)-1] != None:
                            print 'updating row %s: %s=>%s' %(master_row,master_sheet.cell(row=master_row,column=int(x)).value, data_sheet[int(res['index'])][int(y)-1].value)
                            master_sheet.cell(row=master_row,column=int(x)).value = data_sheet[int(res['index'])][int(y)-1].value
                    break



            except Exception,e:
                # print str(e)
                traceback.print_exc()
        if found:
            break

    master_row = master_row + 1

try:
    master_book._add_sheet(master_sheet)
    master_book.save('updates2.xlsx')
except:
    master_book.save('updates2.xlsx')
#
# # sorted_data = list(x for x in [data_sheet.cell(row=y, column=z) for y in range(data_sheet.min_row, data_sheet.max_row) for z in range(data_sheet.min_column, data_sheet.max_column)])
# # sorted_data = list x for x in [data_sheet.cell()]
print found_count


# t1 = threading.Thread(target=t)
# t1.start()
# t1.join()
# row = data_sheet.min_row
# data_list=[]
# while row <= data_sheet.max_row:
#     # print str(row)
#     row_data = []
#     col = data_sheet.min_column
#     while col <= data_sheet.max_column:
#         # print data_sheet.cell(row=row, column=col).value
#         try:
#             print "%s:%s" % (row, col)
#             row_data.append(str(data_sheet.cell(row=row, column=col).value))
#             col=col+1
#         except Exception,e:
#             print str(e)
#             col=col+1
#
#     row = row+1
#     data_list.append(row_data)
#
# master_search_columns = [4,8]
# data_search_columns = [0,3]
# master_row = master_sheet.min_row
# found_count = 0
# # for master_row in master_list:
# while master_row <= master_sheet.max_row:
#     for m in master_search_columns:
#         found = False
#         for d in data_search_columns:
#             data_list = sorted(data_list,key=lambda x: x[d], reverse=False)
#             searchable_data_list = list(x[d] for x in data_list)
#             result = basic_binary_search_with_added_shit(master_sheet.cell(row=master_row, column=m).value,searchable_data_list)
#             found = result['result']
#             # results = binary_search(master_row,data_list, m, d)
#             # found = results["result"]
#             if found:
#                 print "Match found"
#                 found_count = found_count+1
#                 master_sheet.cell(row=master_row, column=20).value = int(data_list[int(result['index'])][2])+int(data_list[int(result['index'])][4])
#
#                 break
#         if found:
#             break
#     master_row = master_row+1
# print str(found_count)
# try:
#     master_book._add_sheet(master_sheet)
#     master_book.save('document.xlsx')
# except:
#     master_book.save('document.xlsx')

#     better_csv = BetterCSV()
#     new_master = []
    # found_count = 0
    # for line in master_copy:
    #     for m in master_search_columns:
    #         found = False
    #         for d in data_search_columns:
    #             data_copy = sorted(data_copy, key=lambda x: x[d], reverse=False)
    #             results = binary_search(line, data_copy,m, d)
    #             found = results["result"]
    #             if found:
    #                 line = update_row(line, data_copy[results["index"]], column_mapping)
    #                 found_count = found_count +1
    #                 break
    #         if found:
    #             break
    #     new_master.append(line)
# master_search_columns = [4,8]
# print "finished loading data"
# #ok so assuming we have our data list we should be able to update the excel
# row = master_sheet.min_row
# data_list=[]
# while row <= master_sheet.max_row:
#     # print "searching row: "+str(row)
#     row_two = data_sheet.min_row
#     data_list=[]
#     while row_two <= data_sheet.max_row:
#         if BetterCSV().search([master_sheet.cell(row=row, column=4).value, master_sheet.cell(row=row, column=8).value], [data_sheet.cell(row=row_two, column=1).value, data_sheet.cell(row=row_two, column=4).value]):
#             print "match found"
#             pass
#         row_two = row_two+1
        # print str(row)
        # row_data = []
        # col = data_sheet.min_column
        # while col <= data_sheet.max_column:
        #     # print data_sheet.cell(row=row, column=col).value
        #     try:
        #         print "%s:%s" % (row, col)
        #         row_data.append(str(data_sheet.cell(row=row, column=col).value))
        #         col=col+1
        #     except Exception,e:
        #         print str(e)
        #         col=col+1
        #
    #don't really feel like doing binary search here right now
    # for r in data_list:
    #     if BetterCSV().search([data_sheet.cell(row=row, column=4).value, data_sheet.cell(row=row, column=8).value], [r[0], r[3]]):
    #         print "Match found"
    #         break
    # pass
    # row = row +1
    # # print str(row)
    # row_data = []
    # col = master_sheet.min_column
    # while col <= master_sheet.max_column:
    #     # print data_sheet.cell(row=row, column=col).value
    #     try:
    #         # print "%s:%s" % (row, col)
    #         row_data.append(str(master_sheet.cell(row=row, column=col).value))
    #         col=col+1
    #     except Exception,e:
    #         print str(e)
    #         col=col+1
    #
    # row = row+1
    # data_list.append(row_data)


# for value in wb.get_sheet_by_name(wb.get_sheet_names()[0]):
#     print "%s: %s" % ("test",value.value)
# sheet_ranges = wb['range names']
# import xlrd
# book = xlrd.open_workbook("price comparison sheet.xlsx")
# print "The number of worksheets is", book.nsheets
# print "Worksheet name(s):", book.sheet_names()
# sh = book.sheet_by_index(0)
# print sh.name, sh.nrows, sh.ncols
# __author__ = 'Andrew'
# from parse_functions import BetterCSV
# from excel_functions import iterate
# # res = BetterCSV().get_lists(["1974,\"6019T88 G03\",,,,,\"Welded Ring\",CF34,,,9431-01,,,,,,14.77,,\"AMS 5754\",\"AMS 5754\",,,,,,,,,,,,"])
# better_csv=BetterCSV()
# master_copy = better_csv.get_lists(better_csv.get_lines(open("data/leah.csv").read()))
# leah = better_csv.get_lists(better_csv.get_lines(open("master_tesdoet.csv").read()))
# master_copy= iterate(master_copy, leah, [3,7],[1,10],{18:3,19:4,20:15,21:16,22:17}, "Master")
# # print "hello "
# # for l in master_copy:
# #     print l
# hs = open("this_is_it_101.csv","w")
# for mline in master_copy:
#     hs.write(",".join(mline)+"\r")
#
# hs.close()